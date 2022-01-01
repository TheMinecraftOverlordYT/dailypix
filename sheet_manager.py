import openpyxl
import datetime


class SheetManager:
    wb = None
    sheet = None

    def __init__(self):
        self.wb = openpyxl.load_workbook("G:/Ketan/PycharmProjects/dailypix/data/records.xlsx")
        need_new = True
        sheets = self.wb.sheetnames
        current_sheet = None
        year = datetime.date.today().strftime("%Y")
        if year in sheets:
            need_new = False
            current_sheet = sheets[sheets.index(year)]

        if need_new:
            current_sheet = self.wb.create_sheet(title=str(year))

        self.sheet = current_sheet

    def get_cell(self):
        month = datetime.date.today().strftime("%m")  # Row
        day = datetime.date.today().strftime("%d")  # Column

        return self.wb[self.sheet.title()].cell(int(day), int(month))

    def check_cell(self):
        cell = self.get_cell()
        return cell.value is None

    def set_cell(self, date, url):
        cell = self.get_cell()
        if cell.value is None:
            cell.value = str(date)[0:10] + " " + url
            self.wb.save("G:/Ketan/PycharmProjects/dailypix/data/records.xlsx")
